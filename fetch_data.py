#!/usr/bin/env python3
"""
Pobiera dzienne ceny commodities z:
  - FRED (Federal Reserve St. Louis) - energia
  - Yahoo Finance (przez pakiet yfinance) - metale, zboza, tropikalne

Wymaga zmiennych srodowiskowych:
  FRED_KEY - klucz API FRED
  EIA_KEY  - klucz API EIA (obecnie nieuzywany, w rezerwie)

Wynik:
  data/prices.json - wszystkie serie w jednym pliku
  data/meta.json   - metadane
"""

import os
import sys
import json
import time
import urllib.request
import urllib.error
import urllib.parse
from pathlib import Path
from datetime import datetime, timezone

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
DATA_DIR.mkdir(exist_ok=True)

FRED_KEY = os.environ.get("FRED_KEY", "").strip()
EIA_KEY = os.environ.get("EIA_KEY", "").strip()

if not FRED_KEY:
    print("BLAD: Ustaw zmienna srodowiskowa FRED_KEY", file=sys.stderr)
    sys.exit(1)

START_DATE = "1995-01-01"  # 30 lat historii; nie wszystkie serie sięgają tak daleko - fetch zwraca co jest dostępne

# ---------------------------------------------------------------------------
# Produkty
#
# source:
#   fred    - pobiera z api.stlouisfed.org
#   yahoo   - pobiera przez yfinance
#   stooq   - pobiera CSV ze stooq.com (dla tickerow ktorych Yahoo nie ma:
#             MATIF Paris, MGEX Spring Wheat, itd)
#   eurostat_elec - detal elektryki (Eurostat nrg_pc_204, polroczne EUR/kWh)
# unit_scale:
#   opcjonalny mnoznik do zastosowania na cenach (np. 0.01 dla ¢ -> $)
#   uzywane zeby wszystkie zboza/kawa mialy jednolite jednostki $
# ---------------------------------------------------------------------------
PRODUCTS = [
    # Energia (FRED) - potwierdzone dzialaja
    dict(id="NG",    source="fred", series="DHHNGSP",
         name="Henry Hub (gaz US)",       unit="$/MMBtu",
         contract_size=10000, contract_unit="MMBtu",
         category="Energia", color="#4fc3f7"),
    dict(id="WTI",   source="fred", series="DCOILWTICO",
         name="WTI Crude (ropa US)",      unit="$/bbl",
         contract_size=1000,  contract_unit="bbl",
         category="Energia", color="#ff8a65"),
    dict(id="BRENT", source="fred", series="DCOILBRENTEU",
         name="Brent Crude (ropa)",       unit="$/bbl",
         contract_size=1000,  contract_unit="bbl",
         category="Energia", color="#ba68c8"),
    dict(id="HEAT",  source="fred", series="DHOILNYH",
         name="Heating Oil / ULSD (NY)",  unit="$/gal",
         contract_size=42000, contract_unit="gal",
         category="Energia", color="#ff5722"),
    dict(id="RBOB",  source="fred", series="DGASUSGULF",
         name="Benzyna (US Gulf spot)",   unit="$/gal",
         contract_size=42000, contract_unit="gal",
         category="Energia", color="#ffb74d"),
    dict(id="GAS_RETAIL_US", source="fred", series="GASREGW",
         name="Benzyna detal USA (avg)",  unit="$/gal",
         contract_size=1, contract_unit="gal",
         category="Energia", color="#ffa000"),
    # Inventories & storage - wskazniki fundamentalne (leading indicators dla cen)
    # UWAGA: FRED nie ma bezposrednich zapasow tygodniowych; usunieto do czasu znalezienia
    # poprawnych series_id (WNGSTUS/WCESTUS zwracaly 400).

    # Metale (Yahoo Finance)
    dict(id="GOLD",     source="yahoo", series="GC=F",
         name="Zloto (COMEX)",            unit="$/oz",
         contract_size=100,   contract_unit="oz",
         category="Metale", color="#ffd54f"),
    dict(id="SILVER",   source="yahoo", series="SI=F",
         name="Srebro (COMEX)",           unit="$/oz",
         contract_size=5000,  contract_unit="oz",
         category="Metale", color="#b0bec5"),
    dict(id="COPPER",   source="yahoo", series="HG=F",
         name="Miedz (COMEX)",            unit="$/lb",
         contract_size=25000, contract_unit="lb",
         category="Metale", color="#d84315"),
    dict(id="PLATINUM", source="yahoo", series="PL=F",
         name="Platyna (NYMEX)",          unit="$/oz",
         contract_size=50,    contract_unit="oz",
         category="Metale", color="#78909c"),
    dict(id="PALLADIUM", source="yahoo", series="PA=F",
         name="Pallad (NYMEX)",           unit="$/oz",
         contract_size=100,   contract_unit="oz",
         category="Metale", color="#8d6e63"),

    # Zboza (Yahoo Finance) - Yahoo daje ceny w centach; przeliczamy na dolary
    dict(id="CORN",    source="yahoo", series="ZC=F", unit_scale=0.01,
         name="Kukurydza (CBOT Chicago)", unit="$/bu",
         contract_size=5000, contract_unit="bu",
         category="Rolne", color="#fbc02d"),
    dict(id="WHEAT",   source="yahoo", series="ZW=F", unit_scale=0.01,
         name="Pszenica SRW (CBOT)",      unit="$/bu",
         contract_size=5000, contract_unit="bu",
         category="Rolne", color="#8d6e63"),
    dict(id="WHEAT_KC", source="yahoo", series="KE=F", unit_scale=0.01,
         name="Pszenica HRW (Kansas)",    unit="$/bu",
         contract_size=5000, contract_unit="bu",
         category="Rolne", color="#a1887f"),
    # WHEAT_MW (MGEX Spring Wheat) usuniete - Stooq nie ma symbolu mw.f, Yahoo tez nie dziala
    # dict(id="WHEAT_MW", source="stooq", series="mw.f", unit_scale=0.01,
    #      name="Pszenica Spring (MGEX)",   unit="$/bu",
    #      contract_size=5000, contract_unit="bu",
    #      category="Rolne", color="#4e342e"),
    dict(id="SOYBEAN", source="yahoo", series="ZS=F", unit_scale=0.01,
         name="Soja (CBOT)",              unit="$/bu",
         contract_size=5000, contract_unit="bu",
         category="Rolne", color="#7cb342"),
    dict(id="SOYMEAL", source="yahoo", series="ZM=F",
         name="Mączka sojowa (CBOT)",     unit="$/ton",
         contract_size=100, contract_unit="short tons",
         category="Rolne", color="#689f38"),
    dict(id="SOYOIL",  source="yahoo", series="ZL=F", unit_scale=0.01,
         name="Olej sojowy (CBOT)",       unit="$/lb",
         contract_size=60000, contract_unit="lb",
         category="Rolne", color="#f9a825"),

    # Polska: retail (detal) z EU Weekly Oil Bulletin - historia tygodniowa od 2005
    # Klucze uzywaja kodow ISO alpha-2 wg struktury pliku WOB (PL_, DE_, EU_, ...)
    dict(id="PL_PB95_WOB", source="wob", series="PL|Euro-Super 95",
         name="Pb95 detal PL (EU WOB)",   unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#4db6ac"),
    dict(id="PL_ON_WOB",   source="wob", series="PL|Automotive Gas Oil",
         name="ON detal PL (EU WOB)",     unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#26a69a"),
    dict(id="PL_LPG_WOB",  source="wob", series="PL|LPG",
         name="LPG detal PL (EU WOB)",    unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#80cbc4"),
    dict(id="PL_HEAT_WOB", source="wob", series="PL|Heating Gas Oil",
         name="Olej opałowy PL (EU WOB)", unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#00897b"),
    # Odniesienie EU-27 i Eurozone
    dict(id="EU_PB95_WOB", source="wob", series="EU|Euro-Super 95",
         name="Pb95 detal UE-27 (WOB)",   unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Europa", color="#9575cd"),
    dict(id="EU_ON_WOB",   source="wob", series="EU|Automotive Gas Oil",
         name="ON detal UE-27 (WOB)",     unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Europa", color="#7e57c2"),
    dict(id="DE_PB95_WOB", source="wob", series="DE|Euro-Super 95",
         name="Pb95 detal Niemcy (WOB)",  unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Europa", color="#b39ddb"),
    dict(id="DE_ON_WOB",   source="wob", series="DE|Automotive Gas Oil",
         name="ON detal Niemcy (WOB)",    unit="€/1000L",
         contract_size=1, contract_unit="L",
         category="Europa", color="#9575cd"),

    # Polska: hurt Orlen (append-only, scraping z cenypaliw.fyi - historia narasta od dzis)
    dict(id="PL_ORLEN_PB95", source="orlen_scrape", series="PB95",
         name="Orlen Pb95 hurt",          unit="zł/L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#ff8f00"),
    dict(id="PL_ORLEN_PB98", source="orlen_scrape", series="PB98",
         name="Orlen Pb98 hurt",          unit="zł/L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#ff6f00"),
    dict(id="PL_ORLEN_ON",   source="orlen_scrape", series="ON",
         name="Orlen ON hurt",            unit="zł/L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#ef6c00"),
    dict(id="PL_ORLEN_EKOTERM", source="orlen_scrape", series="ON_EKOTERM",
         name="Orlen ON Ekoterm hurt",    unit="zł/L",
         contract_size=1, contract_unit="L",
         category="Lokalne PL", color="#e65100"),

    # Dodatkowe kraje EU (WOB) - do porownan miedzynarodowych
    dict(id="FR_PB95_WOB", source="wob", series="FR|Euro-Super 95",
         name="Pb95 detal Francja (WOB)", unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#5c6bc0"),
    dict(id="FR_ON_WOB",   source="wob", series="FR|Automotive Gas Oil",
         name="ON detal Francja (WOB)",   unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#3f51b5"),
    dict(id="ES_PB95_WOB", source="wob", series="ES|Euro-Super 95",
         name="Pb95 detal Hiszpania (WOB)", unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#ec407a"),
    dict(id="ES_ON_WOB",   source="wob", series="ES|Automotive Gas Oil",
         name="ON detal Hiszpania (WOB)", unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#d81b60"),
    dict(id="IT_PB95_WOB", source="wob", series="IT|Euro-Super 95",
         name="Pb95 detal Włochy (WOB)",  unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#66bb6a"),
    dict(id="IT_ON_WOB",   source="wob", series="IT|Automotive Gas Oil",
         name="ON detal Włochy (WOB)",    unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#43a047"),
    dict(id="CZ_PB95_WOB", source="wob", series="CZ|Euro-Super 95",
         name="Pb95 detal Czechy (WOB)",  unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#26a69a"),
    dict(id="CZ_ON_WOB",   source="wob", series="CZ|Automotive Gas Oil",
         name="ON detal Czechy (WOB)",    unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#00897b"),
    dict(id="HU_PB95_WOB", source="wob", series="HU|Euro-Super 95",
         name="Pb95 detal Węgry (WOB)",   unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#ff7043"),
    dict(id="HU_ON_WOB",   source="wob", series="HU|Automotive Gas Oil",
         name="ON detal Węgry (WOB)",     unit="€/1000L",
         contract_size=1, contract_unit="L", category="Europa", color="#f4511e"),

    # Kursy walutowe (do konwersji miedzy jednostkami w rożnych walutach)
    dict(id="FX_EURUSD", source="yahoo", series="EURUSD=X",
         name="EUR/USD (kurs)", unit="USD/EUR",
         contract_size=1, contract_unit="EUR", category="Waluty", color="#42a5f5"),
    dict(id="FX_EURPLN", source="yahoo", series="EURPLN=X",
         name="EUR/PLN (kurs)", unit="PLN/EUR",
         contract_size=1, contract_unit="EUR", category="Waluty", color="#7e57c2"),
    dict(id="FX_USDPLN", source="yahoo", series="PLN=X",
         name="USD/PLN (kurs)", unit="PLN/USD",
         contract_size=1, contract_unit="USD", category="Waluty", color="#26a69a"),
    dict(id="FX_GBPUSD", source="yahoo", series="GBPUSD=X",
         name="GBP/USD (kurs)", unit="USD/GBP",
         contract_size=1, contract_unit="GBP", category="Waluty", color="#5c6bc0"),
    dict(id="FX_USDJPY", source="yahoo", series="JPY=X",
         name="USD/JPY (kurs)", unit="JPY/USD",
         contract_size=1, contract_unit="USD", category="Waluty", color="#ec407a"),
    dict(id="FX_USDCHF", source="yahoo", series="CHF=X",
         name="USD/CHF (kurs)", unit="CHF/USD",
         contract_size=1, contract_unit="USD", category="Waluty", color="#ef5350"),

    # Makro - kluczowe globalne wskazniki wplywajace na surowce
    dict(id="DXY",  source="yahoo", series="DX-Y.NYB",
         name="DXY (Dollar Index)",      unit="pkt",
         contract_size=1000, contract_unit="pkt", category="Makro", color="#7cb342"),
    dict(id="VIX",  source="yahoo", series="^VIX",
         name="VIX (S&P 500 volatility)", unit="pkt",
         contract_size=1000, contract_unit="pkt", category="Makro", color="#ef5350"),
    dict(id="SPX",  source="yahoo", series="^GSPC",
         name="S&P 500 (US equities)",   unit="pkt",
         contract_size=1, contract_unit="pkt", category="Makro", color="#42a5f5"),
    dict(id="BTC",  source="yahoo", series="BTC-USD",
         name="Bitcoin (BTC/USD)",       unit="$/BTC",
         contract_size=1, contract_unit="BTC", category="Makro", color="#ffb300"),
    dict(id="US10Y", source="fred", series="DGS10",
         name="Rentownosc 10Y US",       unit="%",
         contract_size=1, contract_unit="%", category="Makro", color="#5c6bc0"),
    dict(id="US2Y",  source="fred", series="DGS2",
         name="Rentownosc 2Y US",        unit="%",
         contract_size=1, contract_unit="%", category="Makro", color="#7e57c2"),
    dict(id="US10Y2Y", source="fred", series="T10Y2Y",
         name="Spread 10Y-2Y US (recesja)", unit="%",
         contract_size=1, contract_unit="%", category="Makro", color="#d32f2f"),

    # Elektryka hurt (day-ahead) - dzienne, Ember Energy (CSV, bez rejestracji)
    dict(id="PL_POWER", source="ember", series="Poland",
         name="Elektryka hurt Polska",      unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#ffca28"),
    dict(id="DE_POWER", source="ember", series="Germany",
         name="Elektryka hurt Niemcy",      unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#ff9800"),
    dict(id="FR_POWER", source="ember", series="France",
         name="Elektryka hurt Francja",     unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#5c6bc0"),
    dict(id="ES_POWER", source="ember", series="Spain",
         name="Elektryka hurt Hiszpania",   unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#ec407a"),
    dict(id="IT_POWER", source="ember", series="Italy",
         name="Elektryka hurt Włochy",      unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#66bb6a"),
    dict(id="CZ_POWER", source="ember", series="Czechia",
         name="Elektryka hurt Czechy",      unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#26a69a"),
    dict(id="SK_POWER", source="ember", series="Slovakia",
         name="Elektryka hurt Słowacja",    unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#42a5f5"),
    dict(id="HU_POWER", source="ember", series="Hungary",
         name="Elektryka hurt Węgry",       unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#ff7043"),
    dict(id="NL_POWER", source="ember", series="Netherlands",
         name="Elektryka hurt Holandia",    unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#ffb74d"),
    dict(id="GB_POWER", source="ember", series="United Kingdom",
         name="Elektryka hurt UK",          unit="€/MWh",
         contract_size=1, contract_unit="MWh", category="Elektryka", color="#7e57c2"),

    # Elektryka DETAL (gospodarstwa domowe) - Eurostat nrg_pc_204 - TYMCZASOWO WYLACZONE
    # API zwracalo 400 (prawdopodobnie zmiana schematu SDMX). Do naprawy w kolejnym kroku.
    # Zeby przywrocic - odkomentuj i zdiagnozuj URL rucznie.

    # CFTC COT - pozycja NETTO Money Managers (Long - Short) per market, tygodniowo
    # Klasyczny wskaźnik kontrariański - ekstremum netto sygnalizuje odwrócenie
    # UWAGA: CFTC zmienia nazwy rynkow (np. w 2022 dodali suffiksy "-WTI", zmienili exchange).
    # Uzywamy `series_patterns` (lista SUBSTRINGOW) zamiast dokladnej nazwy - fetcher matchuje
    # WSZYSTKIE market names ktore zawieraja KAZDE ze slow kluczowych (AND per pattern grupa).
    # Dane z wielu matchujacych markets sa laczone union po dacie (najnowsza publikacja wygrywa).
    dict(id="COT_WTI_NET",    source="cftc",
         series_patterns=[["CRUDE OIL", "LIGHT SWEET"], ["WTI"]],
         name="COT WTI - Money Mgr netto",  unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#ff8a65"),
    dict(id="COT_BRENT_NET",  source="cftc",
         series_patterns=[["BRENT"]],
         name="COT Brent - Money Mgr netto",unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#ba68c8"),
    dict(id="COT_NG_NET",     source="cftc",
         series_patterns=[["NATURAL GAS"], ["HENRY HUB"], ["NAT GAS"]],
         name="COT NG - Money Mgr netto",   unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#4fc3f7"),
    dict(id="COT_GOLD_NET",   source="cftc",
         series_patterns=[["GOLD"]],
         # UWAGA: to zlapie tez COMEX ETF etc, ale filtrujemy do lucznika: exchange musi zawierac COMEX/CMX/COMMODITY EXCHANGE
         series_exchange_filter=["COMEX", "COMMODITY EXCHANGE"],
         name="COT Gold - Money Mgr netto", unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#ffd54f"),
    dict(id="COT_SILVER_NET", source="cftc",
         series_patterns=[["SILVER"]],
         series_exchange_filter=["COMEX", "COMMODITY EXCHANGE"],
         name="COT Silver - Money Mgr netto", unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#b0bec5"),
    dict(id="COT_COPPER_NET", source="cftc",
         series_patterns=[["COPPER"]],
         series_exchange_filter=["COMEX", "COMMODITY EXCHANGE"],
         name="COT Copper - Money Mgr netto", unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#d84315"),
    dict(id="COT_CORN_NET",   source="cftc",
         series_patterns=[["CORN"]],
         series_exchange_filter=["CHICAGO BOARD OF TRADE", "CBOT"],
         name="COT Corn - Money Mgr netto", unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#fbc02d"),
    dict(id="COT_WHEAT_NET",  source="cftc",
         series_patterns=[["WHEAT-SRW"], ["WHEAT SRW"]],
         series_exchange_filter=["CHICAGO BOARD OF TRADE", "CBOT"],
         name="COT Wheat - Money Mgr netto",unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#8d6e63"),
    dict(id="COT_SOYBEAN_NET",source="cftc",
         series_patterns=[["SOYBEAN"]],
         series_exchange_filter=["CHICAGO BOARD OF TRADE", "CBOT"],
         name="COT Soybean - Money Mgr netto", unit="kontrakty", contract_size=1, contract_unit="kontrakty",
         category="COT", color="#7cb342"),

    # Polska: ceny skupu zbóż - EU Agri-Food Data Portal (tygodniowe/miesieczne)
    dict(id="PL_WHEAT_SKUP",  source="agrifood_cereal", series="Common wheat",
         name="Pszenica skup PL (EU)", unit="€/t",
         contract_size=1, contract_unit="t", category="Lokalne PL", color="#8d6e63"),
    dict(id="PL_CORN_SKUP",   source="agrifood_cereal", series="Feed maize",
         name="Kukurydza skup PL (EU)", unit="€/t",
         contract_size=1, contract_unit="t", category="Lokalne PL", color="#fbc02d"),
    dict(id="PL_RYE_SKUP",    source="agrifood_cereal", series="Rye",
         name="Żyto skup PL (EU)", unit="€/t",
         contract_size=1, contract_unit="t", category="Lokalne PL", color="#a1887f"),
    dict(id="PL_BARLEY_SKUP", source="agrifood_cereal", series="Feed barley",
         name="Jęczmień skup PL (EU)", unit="€/t",
         contract_size=1, contract_unit="t", category="Lokalne PL", color="#c0ca33"),

    # MATIF Paryz (Euronext) - best-effort, Yahoo bywa kaprysny dla EU futures
    dict(id="MATIF_WHEAT", source="stooq", series="ml.f",
         name="Pszenica młynarska (MATIF Paris)", unit="€/t",
         contract_size=50, contract_unit="t",
         category="Rolne", color="#ff7043"),
    dict(id="MATIF_CORN",  source="stooq", series="ema.f",
         name="Kukurydza (MATIF Paris)",  unit="€/t",
         contract_size=50, contract_unit="t",
         category="Rolne", color="#ffab40"),
    dict(id="MATIF_RAPESEED", source="stooq", series="rr.f",
         name="Rzepak (MATIF Paris)",     unit="€/t",
         contract_size=50, contract_unit="t",
         category="Rolne", color="#c0ca33"),

    # Europa - gaz i inne benchmarki europejskie
    dict(id="TTF",   source="yahoo", series="TTF=F",
         name="TTF (gaz EU, Amsterdam)",  unit="€/MWh",
         contract_size=1000, contract_unit="MWh",
         category="Europa", color="#66bb6a"),
    # EUA - EU Emission Allowances (uprawnienia do emisji CO2, ICE Amsterdam)
    # Kluczowe dla marz elektrowni gazowych/weglowych - clean spark/dark spread
    dict(id="EUA",   source="stooq", series="co2.f",
         name="EUA (CO2, EU ETS)",       unit="€/t",
         contract_size=1000, contract_unit="t CO2",
         category="Europa", color="#78909c"),
    dict(id="EU_NG_MO", source="fred", series="PNGASEUUSDM",
         name="Gaz EU (WB, miesieczny)", unit="$/MMBtu",
         contract_size=1, contract_unit="MMBtu",
         category="Europa", color="#81c784"),
    dict(id="COAL_AU", source="fred", series="PCOALAUUSDM",
         name="Wegiel Australia (WB, miesieczny)", unit="$/mt",
         contract_size=1, contract_unit="mt",
         category="Energia", color="#455a64"),
    dict(id="LNG_JP",  source="fred", series="PNGASJPUSDM",
         name="LNG Japonia (WB, miesieczny)", unit="$/MMBtu",
         contract_size=1, contract_unit="MMBtu",
         category="Energia", color="#e57373"),

    # Tropikalne (Yahoo Finance)
    dict(id="COFFEE",  source="yahoo", series="KC=F", unit_scale=0.01,
         name="Kawa Arabica (ICE)",       unit="$/lb",
         contract_size=37500, contract_unit="lb",
         category="Rolne", color="#6d4c41"),
    dict(id="COCOA",   source="yahoo", series="CC=F",
         name="Kakao (ICE)",              unit="$/mt",
         contract_size=10,   contract_unit="mt",
         category="Rolne", color="#5d4037"),
    dict(id="SUGAR",   source="yahoo", series="SB=F", unit_scale=0.01,
         name="Cukier (ICE 11)",          unit="$/lb",
         contract_size=112000, contract_unit="lb",
         category="Rolne", color="#f8bbd0"),
    dict(id="COTTON",  source="yahoo", series="CT=F", unit_scale=0.01,
         name="Bawelna (ICE)",            unit="$/lb",
         contract_size=50000, contract_unit="lb",
         category="Rolne", color="#f5f5dc"),
    dict(id="OJ",      source="yahoo", series="OJ=F", unit_scale=0.01,
         name="Sok pomaranczowy (ICE)",   unit="$/lb",
         contract_size=15000, contract_unit="lb",
         category="Rolne", color="#ff9800"),
    # Zywiec (Yahoo, CME)
    dict(id="LIVE_CATTLE", source="yahoo", series="LE=F", unit_scale=0.01,
         name="Bydlo zywe (CME)",         unit="$/lb",
         contract_size=40000, contract_unit="lb",
         category="Rolne", color="#8d6e63"),
    dict(id="LEAN_HOGS",   source="yahoo", series="HE=F", unit_scale=0.01,
         name="Trzoda chlewna (CME)",     unit="$/lb",
         contract_size=40000, contract_unit="lb",
         category="Rolne", color="#e91e63"),
]


def fetch_fred(series_id: str) -> list:
    """Zwraca liste {date, value} dla serii FRED."""
    url = (
        "https://api.stlouisfed.org/fred/series/observations"
        f"?series_id={series_id}&api_key={FRED_KEY}&file_type=json"
        f"&observation_start={START_DATE}"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "energy-analytics/1.0"})
    with urllib.request.urlopen(req, timeout=60) as r:
        data = json.load(r)
    obs = []
    for o in data.get("observations", []):
        v = o.get("value", "")
        if v and v != ".":
            try:
                obs.append({"date": o["date"], "value": float(v)})
            except ValueError:
                pass
    return obs


WOB_URL = ("https://energy.ec.europa.eu/document/download/"
           "906e60ca-8b6a-44e7-8589-652854d2fd3f_en"
           "?filename=Weekly_Oil_Bulletin_Prices_History_maticni_4web.xlsx")
_WOB_CACHE = None  # cache pobranego XLSX - jeden download na cala petle produktow

def _download_wob_xlsx():
    """Downloaduje WOB XLSX raz i cache'uje surowe bajty."""
    global _WOB_CACHE
    if _WOB_CACHE is not None:
        return _WOB_CACHE
    print(f"  [WOB] pobieram XLSX z ec.europa.eu ...", flush=True)
    req = urllib.request.Request(WOB_URL, headers={"User-Agent": "Mozilla/5.0 (energy-analytics)"})
    with urllib.request.urlopen(req, timeout=120) as r:
        _WOB_CACHE = r.read()
    print(f"  [WOB] pobrano {len(_WOB_CACHE)/1024/1024:.1f} MB", flush=True)
    return _WOB_CACHE


_WOB_PARSED = None

def _parse_wob_workbook():
    """Parsuje XLSX WOB (wielopoziomowe naglowki: kraj + produkt).
    Zwraca slownik: {(country, product): [(date_iso, price), ...]}.
    Wynik cache'owany globalnie - pobierany raz na uruchomienie skryptu.
    """
    global _WOB_PARSED
    if _WOB_PARSED is not None:
        return _WOB_PARSED

    import openpyxl
    from io import BytesIO
    from datetime import datetime as _dt

    raw = _download_wob_xlsx()
    wb = openpyxl.load_workbook(BytesIO(raw), read_only=False, data_only=True)
    print(f"  [WOB] arkusze: {wb.sheetnames}", flush=True)

    if 'Prices with taxes' not in wb.sheetnames:
        print(f"  [WOB] BLAD: brak arkusza 'Prices with taxes'", flush=True)
        _WOB_PARSED = {}
        return _WOB_PARSED
    ws = wb['Prices with taxes']

    # Merged cells: rozprowadz wartosc top-left po calym merge
    merged_values = {}
    for mrange in ws.merged_cells.ranges:
        top = ws.cell(row=mrange.min_row, column=mrange.min_col).value
        for r in range(mrange.min_row, mrange.max_row + 1):
            for c in range(mrange.min_col, mrange.max_col + 1):
                merged_values[(r, c)] = top

    def cell(r, c):
        v = merged_values.get((r, c))
        if v is not None: return v
        return ws.cell(row=r, column=c).value

    print(f"  [WOB] wymiary: {ws.max_row} wierszy x {ws.max_column} kolumn", flush=True)

    def normalize_product(s):
        if not s: return None
        s = str(s).strip().lower()
        if 'euro-super 95' in s or 'euro super 95' in s: return 'Euro-Super 95'
        if 'gas oil automob' in s or 'diesel' in s:      return 'Automotive Gas Oil'
        if 'gas oil de cha' in s or 'heating' in s:      return 'Heating Gas Oil'
        if 'gpl' in s or 'lpg' in s:                     return 'LPG'
        if 'fuel oil -schw' in s.replace(' ', ''):       return 'Residual Fuel Oil (low S)'
        if 'fuel oil - sch' in s or 'fuel oil -sch' in s: return 'Residual Fuel Oil (high S)'
        return None

    product_by_col = {}
    for c in range(2, ws.max_column + 1):
        v = cell(2, c)
        prod = normalize_product(v)
        if prod:
            product_by_col[c] = prod

    print(f"  [WOB] kolumn z produktami: {len(product_by_col)}", flush=True)

    first_data_row = None
    for r in range(3, 10):
        v1 = cell(r, 1)
        if hasattr(v1, 'strftime'):
            first_data_row = r
            break
        if isinstance(v1, str):
            try:
                _dt.strptime(v1[:10], '%Y-%m-%d')
                first_data_row = r
                break
            except: pass
    if first_data_row is None:
        first_data_row = 4
    print(f"  [WOB] pierwsze dane w wierszu {first_data_row}", flush=True)

    country_by_col = {}
    for c in range(1, ws.max_column + 1):
        v = cell(first_data_row, c)
        if isinstance(v, str) and v.endswith('_') and 1 <= len(v) <= 5:
            country_by_col[c] = v.rstrip('_')

    print(f"  [WOB] wykryte kolumny markerow krajow: {len(country_by_col)}", flush=True)
    print(f"  [WOB] kody krajow: {sorted(set(country_by_col.values()))}", flush=True)

    sorted_country_cols = sorted(country_by_col.keys())
    col_to_cp = {}
    for prod_col, product_name in product_by_col.items():
        cc_code = None
        for cc in sorted_country_cols:
            if cc <= prod_col:
                cc_code = country_by_col[cc]
            else:
                break
        if cc_code:
            col_to_cp[prod_col] = (cc_code, product_name)

    print(f"  [WOB] zmapowanych par kraj+produkt: {len(col_to_cp)}", flush=True)
    seen_countries = set(cp[0] for cp in col_to_cp.values())
    print(f"  [WOB] kraje w mapie: {sorted(seen_countries)}", flush=True)

    results = {}
    n_scanned = 0
    n_valid = 0
    for r in range(first_data_row, ws.max_row + 1):
        n_scanned += 1
        date_val = cell(r, 1)
        date_str = None
        if hasattr(date_val, 'strftime'):
            date_str = date_val.strftime('%Y-%m-%d')
        elif isinstance(date_val, str):
            try:
                date_str = _dt.strptime(date_val[:10], '%Y-%m-%d').strftime('%Y-%m-%d')
            except: pass
        if not date_str:
            continue
        for col, (country, product) in col_to_cp.items():
            val = cell(r, col)
            if isinstance(val, (int, float)) and 0 < val < 100000:
                results.setdefault((country, product), []).append((date_str, float(val)))
                n_valid += 1

    print(f"  [WOB] przeskanowano {n_scanned} wierszy danych, {n_valid} valid values", flush=True)
    for k in [('PL', 'Euro-Super 95'), ('PL', 'Automotive Gas Oil'), ('PL', 'LPG'), ('EU', 'Euro-Super 95'), ('DE', 'Euro-Super 95')]:
        vals = results.get(k, [])
        if vals:
            print(f"  [WOB] check {k[0]}|{k[1]}: {len(vals)} obs, ostatnia {vals[-1]}", flush=True)
        else:
            print(f"  [WOB] check {k[0]}|{k[1]}: BRAK", flush=True)

    _WOB_PARSED = results
    return results


def fetch_wob(series_key: str) -> list:
    """Zwraca liste {date, value} dla klucza 'Country|Product'."""
    parsed = _parse_wob_workbook()
    if '|' not in series_key:
        return []
    country, product = series_key.split('|', 1)
    obs_tuples = parsed.get((country, product), [])
    if not obs_tuples:
        for (c, p), vals in parsed.items():
            if c.lower() == country.lower() and p.lower() == product.lower():
                obs_tuples = vals
                break
    seen = {}
    for d, v in obs_tuples:
        seen[d] = v
    obs = [{"date": d, "value": v} for d, v in sorted(seen.items())]
    return obs


_AGRIFOOD_CACHE = None

def _fetch_agrifood_cereal_pl():
    """Pobiera tygodniowe ceny cereal z EU Agri-Food Data Portal dla Polski."""
    global _AGRIFOOD_CACHE
    if _AGRIFOOD_CACHE is not None:
        return _AGRIFOOD_CACHE
    from datetime import datetime as _dt3

    url = "https://api.tech.ec.europa.eu/agrifood/api/cereal/prices?memberStateCodes=PL&limit=10000"
    print(f"  [AGRIFOOD] pobieram ceny zbóż PL z EU Agri Portal ...", flush=True)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (energy-analytics)", "Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=60) as r:
            data = json.load(r)
    except Exception as e:
        print(f"  [AGRIFOOD] BLAD: {e}", flush=True)
        _AGRIFOOD_CACHE = {}
        return _AGRIFOOD_CACHE

    items = data if isinstance(data, list) else (data.get('items') if isinstance(data, dict) else [])
    if not items:
        print(f"  [AGRIFOOD] brak items w odpowiedzi", flush=True)
        _AGRIFOOD_CACHE = {}
        return _AGRIFOOD_CACHE
    print(f"  [AGRIFOOD] wpisow: {len(items)}", flush=True)

    def parse_price(s):
        if s is None: return None
        s = str(s).replace('€', '').replace('EUR', '').replace(' ', '').replace('\xa0', '').strip()
        if ',' in s and '.' in s:
            s = s.replace('.', '').replace(',', '.')
        else:
            s = s.replace(',', '.')
        try: return float(s)
        except ValueError: return None

    def parse_date(s):
        if not s: return None
        s = str(s).strip()
        for fmt in ('%d/%m/%Y', '%Y-%m-%d', '%d.%m.%Y', '%d-%m-%Y'):
            try: return _dt3.strptime(s[:10], fmt).strftime('%Y-%m-%d')
            except: pass
        return None

    result = {}
    n_ok = 0
    unique_products = set()
    for row in items:
        if not isinstance(row, dict): continue
        product = row.get('productName') or row.get('product') or row.get('name')
        if not product: continue
        unique_products.add(product)
        date_str = parse_date(row.get('beginDate') or row.get('endDate') or row.get('date') or row.get('referencePeriod'))
        if not date_str: continue
        price = parse_price(row.get('price'))
        if price is None or price <= 0: continue
        result.setdefault(product, []).append({'date': date_str, 'value': price})
        n_ok += 1

    for prod in list(result.keys()):
        seen = {}
        for o in result[prod]:
            seen[o['date']] = o['value']
        result[prod] = [{'date': d, 'value': v} for d, v in sorted(seen.items())]

    print(f"  [AGRIFOOD] wczytano {n_ok} valid wierszy", flush=True)
    print(f"  [AGRIFOOD] unikalne productName ({len(unique_products)}): {sorted(unique_products)}", flush=True)
    _AGRIFOOD_CACHE = result
    return _AGRIFOOD_CACHE


AGRIFOOD_ALIASES = {
    "Common wheat":  ["common wheat", "soft wheat", "wheat", "blt", "blé tendre", "pszenica"],
    "Feed maize":    ["feed maize", "maize", "corn", "mai", "kukurydza"],
    "Rye":           ["rye", "seg", "seigle", "żyto", "zyto"],
    "Feed barley":   ["feed barley", "barley", "org", "orge", "jęczmień", "jeczmien"],
}

def fetch_agrifood_cereal(product_name: str) -> list:
    """Zwraca liste {date, value} dla podanego produktu."""
    parsed = _fetch_agrifood_cereal_pl()
    if not parsed:
        return []
    if product_name in parsed:
        return parsed[product_name]
    aliases = [product_name.lower()] + [a.lower() for a in AGRIFOOD_ALIASES.get(product_name, [])]
    matched_keys = []
    for k in parsed.keys():
        kl = k.lower()
        if any(a == kl or a in kl or kl in a for a in aliases):
            matched_keys.append(k)
    if not matched_keys:
        return []
    combined = {}
    for k in matched_keys:
        for o in parsed[k]:
            combined[o["date"]] = o["value"]
    return [{"date": d, "value": v} for d, v in sorted(combined.items())]


_CFTC_CACHE = None

def _fetch_cftc_all():
    """Pobiera dane CFTC COT."""
    global _CFTC_CACHE
    if _CFTC_CACHE is not None:
        return _CFTC_CACHE

    all_keywords = set()
    for p in PRODUCTS:
        if p["source"] != "cftc": continue
        for pat_group in p.get("series_patterns", []):
            for kw in pat_group:
                all_keywords.add(kw.upper())
    if not all_keywords:
        _CFTC_CACHE = {}
        return _CFTC_CACHE

    like_clauses = " OR ".join(f"upper(market_and_exchange_names) like '%{kw}%'" for kw in sorted(all_keywords))
    where = f"({like_clauses}) AND report_date_as_yyyy_mm_dd > '2010-01-01T00:00:00'"
    params = urllib.parse.urlencode({
        "$select": "market_and_exchange_names,report_date_as_yyyy_mm_dd,m_money_positions_long_all,m_money_positions_short_all",
        "$where": where,
        "$order": "report_date_as_yyyy_mm_dd",
        "$limit": "200000",
    })
    url = "https://publicreporting.cftc.gov/resource/72hh-3qpy.json?" + params

    print(f"  [CFTC] pobieram COT z {len(all_keywords)} keyword-patterns (broad LIKE) ...", flush=True)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "energy-analytics/1.0"})
        with urllib.request.urlopen(req, timeout=180) as r:
            rows = json.load(r)
        print(f"  [CFTC] pobrano {len(rows)} wierszy", flush=True)
    except Exception as e:
        print(f"  [CFTC] BLAD: {e}", flush=True)
        _CFTC_CACHE = {}
        return _CFTC_CACHE

    result = {}
    counts_per_market = {}
    latest_per_market = {}
    for row in rows:
        market = row.get("market_and_exchange_names", "").strip()
        raw_date = row.get("report_date_as_yyyy_mm_dd", "") or ""
        date_str = raw_date[:10] if raw_date else ""
        try:
            long_p = float(row.get("m_money_positions_long_all") or 0)
            short_p = float(row.get("m_money_positions_short_all") or 0)
        except (ValueError, TypeError):
            continue
        if not market or not date_str: continue
        net = long_p - short_p
        result.setdefault(market, []).append({"date": date_str, "value": net})
        counts_per_market[market] = counts_per_market.get(market, 0) + 1
        if date_str > latest_per_market.get(market, ""):
            latest_per_market[market] = date_str

    for market in list(result.keys()):
        seen = {}
        for o in result[market]:
            seen[o["date"]] = o["value"]
        result[market] = [{"date": d, "value": v} for d, v in sorted(seen.items())]

    for m, c in sorted(counts_per_market.items(), key=lambda x: -x[1])[:40]:
        print(f"  [CFTC] {c:>5}w, ost {latest_per_market.get(m,'-')}: {m[:75]}", flush=True)

    _CFTC_CACHE = result
    return _CFTC_CACHE


def fetch_cftc_patterns(patterns, exchange_filter=None) -> list:
    """Zwraca liste {date, value} dla rynkow matchujacych patterns."""
    cache = _fetch_cftc_all()
    exch_filter = [e.upper() for e in (exchange_filter or [])]
    matched_markets = []
    for market in cache.keys():
        mu = market.upper()
        group_match = any(all(kw.upper() in mu for kw in group) for group in patterns)
        if not group_match: continue
        if exch_filter and not any(ef in mu for ef in exch_filter): continue
        matched_markets.append(market)

    if not matched_markets:
        print(f"  [CFTC-match] BRAK dopasowania dla patternow {patterns} (exch={exchange_filter})", flush=True)
        return []
    print(f"  [CFTC-match] {len(matched_markets)} rynkow: {[m[:60] for m in matched_markets]}", flush=True)
    combined = {}
    for m in matched_markets:
        for o in cache[m]:
            combined[o["date"]] = o["value"]
    return [{"date": d, "value": v} for d, v in sorted(combined.items())]


def fetch_cftc(market_name: str, alts=None) -> list:
    return fetch_cftc_patterns([[market_name]] + [[a] for a in (alts or [])])


_EMBER_CACHE = None
EMBER_URL = "https://files.ember-energy.org/public-downloads/price/outputs/european_wholesale_electricity_price_data_daily.csv"

def _parse_ember():
    """Parsuje CSV z Ember Energy."""
    global _EMBER_CACHE
    if _EMBER_CACHE is not None:
        return _EMBER_CACHE
    print(f"  [EMBER] pobieram CSV z Ember Energy ...", flush=True)
    try:
        req = urllib.request.Request(EMBER_URL, headers={"User-Agent": "Mozilla/5.0 (energy-analytics)"})
        with urllib.request.urlopen(req, timeout=60) as r:
            raw = r.read().decode('utf-8', errors='replace')
        print(f"  [EMBER] pobrano {len(raw)/1024:.1f} KB", flush=True)
    except Exception as e:
        print(f"  [EMBER] BLAD: {e}", flush=True)
        _EMBER_CACHE = {}
        return _EMBER_CACHE

    lines = raw.strip().split('\n')
    if len(lines) < 2:
        _EMBER_CACHE = {}
        return _EMBER_CACHE

    header = [h.strip().strip('"') for h in lines[0].split(',')]
    print(f"  [EMBER] header: {header}", flush=True)
    col_country = col_date = col_price = None
    for i, h in enumerate(header):
        hl = h.lower()
        if col_country is None and ('country' in hl and 'iso' not in hl and 'code' not in hl):
            col_country = i
        if col_date is None and 'date' in hl:
            col_date = i
        if col_price is None and ('price' in hl or 'eur' in hl):
            col_price = i

    if col_country is None or col_date is None or col_price is None:
        print(f"  [EMBER] BLAD: nie znaleziono kolumn (country={col_country}, date={col_date}, price={col_price})", flush=True)
        _EMBER_CACHE = {}
        return _EMBER_CACHE

    print(f"  [EMBER] kolumny: country={col_country}, date={col_date}, price={col_price}", flush=True)

    results = {}
    n_ok = 0
    for line in lines[1:]:
        parts = [p.strip().strip('"') for p in line.split(',')]
        if len(parts) <= max(col_country, col_date, col_price):
            continue
        country = parts[col_country]
        date_s = parts[col_date][:10]
        try:
            price = float(parts[col_price])
        except ValueError:
            continue
        try:
            datetime.strptime(date_s, '%Y-%m-%d')
        except ValueError:
            continue
        results.setdefault(country, []).append({'date': date_s, 'value': price})
        n_ok += 1

    for country in results:
        results[country].sort(key=lambda o: o['date'])

    print(f"  [EMBER] wczytano {n_ok} obserwacji, {len(results)} krajow", flush=True)
    for c in ['Poland', 'Germany', 'France']:
        v = results.get(c, [])
        if v:
            print(f"  [EMBER] {c}: {len(v)} obs, ostatnia {v[-1]}", flush=True)
        else:
            print(f"  [EMBER] {c}: BRAK", flush=True)

    _EMBER_CACHE = results
    return results


def fetch_ember(country: str) -> list:
    """Zwraca liste {date, value} dla podanego kraju."""
    return _parse_ember().get(country, [])


_ORLEN_CACHE = None

def _scrape_orlen_current():
    """Scrapuje aktualne hurtowe ceny Orlen z cenypaliw.fyi."""
    global _ORLEN_CACHE
    if _ORLEN_CACHE is not None:
        return _ORLEN_CACHE
    import re
    url = "https://cenypaliw.fyi/"
    print(f"  [ORLEN] scraping {url} ...", flush=True)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0 (energy-analytics)"})
        with urllib.request.urlopen(req, timeout=30) as r:
            html = r.read().decode('utf-8', errors='replace')
        print(f"  [ORLEN] pobrano {len(html)/1024:.1f} KB HTML", flush=True)
    except Exception as e:
        print(f"  [ORLEN] BLAD pobierania: {e}", flush=True)
        _ORLEN_CACHE = {}
        return _ORLEN_CACHE

    result = {}
    section_m = re.search(
        r'(?:Tabela\s+hurtowych\s+cen\s+paliw|hurtowe\s+ceny\s+paliw)(.{0,5000})',
        html, re.DOTALL | re.IGNORECASE
    )
    section = section_m.group(1) if section_m else html
    print(f"  [ORLEN] szukam w sekcji {len(section)} znakow (znaleziona sekcja hurt: {bool(section_m)})", flush=True)

    patterns = {
        'PB95': [
            r'PB\s*95[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*PLN',
            r'PB\s*95[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*z[łl]',
        ],
        'PB98': [
            r'PB\s*98[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*PLN',
            r'PB\s*98[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*z[łl]',
        ],
        'ON': [
            r'(?<![a-zA-Z])ON(?!\s*Ekoterm)\s*(?:\(Diesel\))?[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*PLN',
            r'(?<![a-zA-Z])ON(?!\s*Ekoterm)\s*(?:\(Diesel\))?[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*z[łl]',
        ],
        'ON_EKOTERM': [
            r'ON\s*Ekoterm[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*PLN',
            r'Ekoterm[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*PLN',
            r'ON\s*Ekoterm[^0-9]{1,30}?(\d[\d,\.]{1,6})\s*z[łl]',
        ],
    }
    for key, pats in patterns.items():
        for pat in pats:
            m = re.search(pat, section, re.IGNORECASE)
            if m:
                try:
                    v = float(m.group(1).replace(',', '.'))
                    if 3.0 < v < 9.0:
                        result[key] = v
                        break
                except: pass

    if 'ON_EKOTERM' not in result:
        ekoterm_patterns = [
            r'(?:ON\s*)?Ekoterm[^0-9]{1,60}?(\d[\d,\.]{1,6})\s*(?:PLN|z[łl])',
            r'olej\s+opa[łl]owy[^0-9]{1,80}?(\d[\d,\.]{1,6})\s*(?:PLN|z[łl])',
            r'Ekoterm\s+Plus[^0-9]{1,60}?(\d[\d,\.]{1,6})',
        ]
        for pat in ekoterm_patterns:
            for m in re.finditer(pat, html, re.IGNORECASE):
                try:
                    v = float(m.group(1).replace(',', '.'))
                    if 2.0 < v < 8.0:
                        result['ON_EKOTERM'] = v
                        break
                except: pass
            if 'ON_EKOTERM' in result: break

    print(f"  [ORLEN] znalezione ceny: {result}", flush=True)
    _ORLEN_CACHE = result
    return result


def fetch_orlen_append(product_key: str, existing_obs: list) -> list:
    """Dopisuje/nadpisuje dzisiejsza cene do istniejacej historii Orlen."""
    from datetime import datetime as _dt2, timezone as _tz2
    today = _dt2.now(_tz2.utc).strftime('%Y-%m-%d')
    prices = _scrape_orlen_current()
    today_val = prices.get(product_key)
    obs = [o for o in (existing_obs or []) if o.get('date') != today]
    if today_val is not None:
        obs.append({'date': today, 'value': today_val})
        obs.sort(key=lambda o: o['date'])
    return obs


_EUROSTAT_ELEC_CACHE = None

def _fetch_eurostat_elec_all():
    """Pobiera detaliczne ceny elektryki z Eurostat nrg_pc_204."""
    global _EUROSTAT_ELEC_CACHE
    if _EUROSTAT_ELEC_CACHE is not None:
        return _EUROSTAT_ELEC_CACHE

    geos = [p["series"] for p in PRODUCTS if p["source"] == "eurostat_elec"]
    if not geos:
        _EUROSTAT_ELEC_CACHE = {}
        return _EUROSTAT_ELEC_CACHE

    params_list = [
        ("format", "JSON"),
        ("lang", "EN"),
        ("product", "6000"),
        ("nrg_cons", "KWH2500-4999"),
        ("unit", "KWH"),
        ("tax", "I_TAX"),
        ("currency", "EUR"),
    ]
    for g in geos:
        params_list.append(("geo", g))
    query = urllib.parse.urlencode(params_list)
    url = f"https://ec.europa.eu/eurostat/api/dissemination/statistics/1.0/data/nrg_pc_204?{query}"

    print(f"  [EUROSTAT_ELEC] pobieram detal elektryki dla {len(geos)} krajow ...", flush=True)
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "energy-analytics/1.0"})
        with urllib.request.urlopen(req, timeout=90) as r:
            data = json.load(r)
    except Exception as e:
        print(f"  [EUROSTAT_ELEC] BLAD: {e}", flush=True)
        _EUROSTAT_ELEC_CACHE = {}
        return _EUROSTAT_ELEC_CACHE

    dim_ids = data.get("id", [])
    dim_sizes = data.get("size", [])
    dims = data.get("dimension", {})
    values = data.get("value", {})
    if not dim_ids or not values:
        print(f"  [EUROSTAT_ELEC] pusta odpowiedz", flush=True)
        _EUROSTAT_ELEC_CACHE = {}
        return _EUROSTAT_ELEC_CACHE

    dim_index = {}
    for did in dim_ids:
        cat = dims.get(did, {}).get("category", {})
        idx_map = cat.get("index", {})
        pos_to_id = [None] * len(idx_map)
        for cid, pos in idx_map.items():
            pos_to_id[pos] = cid
        dim_index[did] = pos_to_id

    def unflatten(flat_idx):
        out = {}
        for i in range(len(dim_ids) - 1, -1, -1):
            sz = dim_sizes[i]
            out[dim_ids[i]] = dim_index[dim_ids[i]][flat_idx % sz]
            flat_idx //= sz
        return out

    def period_to_date(t):
        t = str(t)
        if "-S1" in t: return t[:4] + "-01-01"
        if "-S2" in t: return t[:4] + "-07-01"
        if len(t) == 4 and t.isdigit(): return t + "-01-01"
        return None

    result = {}
    for flat_key, v in values.items():
        try:
            fk = int(flat_key)
            fv = float(v)
        except (ValueError, TypeError):
            continue
        coord = unflatten(fk)
        geo = coord.get("geo")
        period = coord.get("time")
        date_str = period_to_date(period)
        if not geo or not date_str: continue
        result.setdefault(geo, []).append({"date": date_str, "value": round(fv, 5)})

    for g in list(result.keys()):
        seen = {}
        for o in result[g]:
            seen[o["date"]] = o["value"]
        result[g] = [{"date": d, "value": v} for d, v in sorted(seen.items())]
        print(f"  [EUROSTAT_ELEC] {g}: {len(result[g])} punktow", flush=True)

    _EUROSTAT_ELEC_CACHE = result
    return _EUROSTAT_ELEC_CACHE


def fetch_eurostat_elec(geo_code: str) -> list:
    return _fetch_eurostat_elec_all().get(geo_code, [])


def fetch_stooq(symbol: str, unit_scale: float = 1.0) -> list:
    """Pobiera dzienne dane ze Stooq (CSV)."""
    import urllib.request
    sym = symbol.lower()
    url = f"https://stooq.com/q/d/l/?s={urllib.parse.quote(sym)}&i=d"
    try:
        req = urllib.request.Request(url, headers={"User-Agent": "energy-analytics/1.0"})
        with urllib.request.urlopen(req, timeout=90) as r:
            raw = r.read().decode("utf-8", errors="replace")
    except Exception as e:
        print(f"  [STOOQ {symbol}] BLAD pobrania: {e}", flush=True)
        return []
    if not raw or "Date" not in raw.split("\n", 1)[0]:
        print(f"  [STOOQ {symbol}] pusta odpowiedz lub blad formatu", flush=True)
        return []
    obs = []
    import csv, io
    reader = csv.DictReader(io.StringIO(raw))
    for row in reader:
        date_str = (row.get("Date") or "").strip()
        close_str = (row.get("Close") or "").strip()
        if not date_str or not close_str: continue
        try:
            v = float(close_str) * unit_scale
        except ValueError:
            continue
        if date_str < START_DATE: continue
        obs.append({"date": date_str, "value": round(v, 4)})
    obs.sort(key=lambda o: o["date"])
    return obs


def fetch_yahoo(ticker: str, unit_scale: float = 1.0) -> list:
    """Zwraca liste {date, value} z Yahoo Finance przez yfinance."""
    try:
        import yfinance as yf
    except ImportError:
        raise RuntimeError("Brakuje pakietu yfinance - dodaj do requirements.txt")

    t = yf.Ticker(ticker)
    df = t.history(start=START_DATE, auto_adjust=False, actions=False)
    if df is None or df.empty:
        return []
    obs = []
    for ts, row in df.iterrows():
        close = row.get("Close")
        if close is None:
            continue
        try:
            v = float(close) * unit_scale
        except (TypeError, ValueError):
            continue
        if v != v: continue  # pomija NaN (NaN != NaN to zawsze True)
        # ts moze byc pd.Timestamp; wez ISO date
        date_str = ts.strftime("%Y-%m-%d") if hasattr(ts, "strftime") else str(ts)[:10]
        obs.append({"date": date_str, "value": round(v, 4)})
    return obs


def _load_existing_prices():
    """Zwraca istniejacy prices.json (jesli jest) jako slownik products {pid: [obs]}."""
    fp = DATA_DIR / "prices.json"
    if not fp.exists():
        return {}
    try:
        with fp.open() as f:
            data = json.load(f)
        return data.get('products', {})
    except Exception as e:
        print(f"[WARN] Nie moge odczytac istniejacego prices.json: {e}", flush=True)
        return {}


def _merge_history(existing_obs, fresh_obs):
    """Laczy istniejaca historie z nowymi danymi."""
    if not existing_obs and not fresh_obs: return []
    if not existing_obs: return list(fresh_obs)
    if not fresh_obs: return list(existing_obs)
    combined = {o['date']: o['value'] for o in existing_obs}
    for o in fresh_obs:
        combined[o['date']] = o['value']
    return [{'date': d, 'value': v} for d, v in sorted(combined.items())]


def main():
    print(f"Start: {datetime.now(timezone.utc).isoformat()}")
    print(f"Zakres historii: od {START_DATE}")

    existing = _load_existing_prices()
    print(f"Wczytano istniejaca historie: {len(existing)} produktow", flush=True)

    products_out = {}
    meta_out = []
    failed = []

    for p in PRODUCTS:
        pid = p["id"]
        scale = p.get("unit_scale", 1.0)
        scale_note = f" (skala x{scale})" if scale != 1.0 else ""
        series_label = p.get('series') or ('|'.join('+'.join(g) for g in p.get('series_patterns', [])) or '(patterns)')
        print(f"\n[{pid}] {p['source']}:{series_label}{scale_note} ...", flush=True)
        try:
            if p["source"] == "fred":
                obs = fetch_fred(p["series"])
            elif p["source"] == "yahoo":
                obs = fetch_yahoo(p["series"], scale)
            elif p["source"] == "stooq":
                obs = fetch_stooq(p["series"], scale)
            elif p["source"] == "wob":
                obs = fetch_wob(p["series"])
            elif p["source"] == "orlen_scrape":
                obs = fetch_orlen_append(p["series"], existing.get(pid, []))
            elif p["source"] == "ember":
                obs = fetch_ember(p["series"])
            elif p["source"] == "eurostat_elec":
                obs = fetch_eurostat_elec(p["series"])
            elif p["source"] == "cftc":
                if p.get("series_patterns"):
                    obs = fetch_cftc_patterns(p["series_patterns"], p.get("series_exchange_filter"))
                else:
                    obs = fetch_cftc(p.get("series", ""), p.get("series_alt"))
            elif p["source"] == "agrifood_cereal":
                obs = fetch_agrifood_cereal(p["series"])
            else:
                print(f"  [SKIP] nieznane zrodlo {p['source']}")
                continue

            existing_obs = existing.get(pid, [])
            merged = _merge_history(existing_obs, obs)
            if not merged:
                print(f"  [WARN] brak obserwacji (i brak historii)")
                failed.append(pid)
                continue
            if not obs and existing_obs:
                print(f"  [WARN] fresh pusty - zachowano {len(existing_obs)} historycznych obs")
            elif obs and len(existing_obs) > len(obs) + 5:
                print(f"  [INFO] fresh={len(obs)} < historia={len(existing_obs)}, merged={len(merged)}")

            products_out[pid] = merged
            first, last = merged[0], merged[-1]
            print(f"  [OK] {len(merged)} obs (fresh {len(obs)}), {first['date']} -> {last['date']}, "
                  f"ostatnia = {last['value']} {p['unit']}")
            meta_out.append({
                "id": pid, "name": p["name"], "unit": p["unit"],
                "contract_size": p["contract_size"], "contract_unit": p["contract_unit"],
                "category": p["category"], "color": p["color"],
                "source": p["source"], "series": p["series"],
                "first_date": first["date"], "last_date": last["date"],
                "n_observations": len(merged),
            })
        except urllib.error.HTTPError as e:
            print(f"  [BLAD HTTP {e.code}] {e.reason}")
            failed.append(pid)
            existing_obs = existing.get(pid, [])
            if existing_obs:
                products_out[pid] = existing_obs
                print(f"  [SAVE] zachowano {len(existing_obs)} historycznych obs")
        except Exception as e:
            print(f"  [BLAD] {type(e).__name__}: {e}")
            failed.append(pid)
            existing_obs = existing.get(pid, [])
            if existing_obs:
                products_out[pid] = existing_obs
                print(f"  [SAVE] zachowano {len(existing_obs)} historycznych obs")
        time.sleep(0.3)

    prices_file = DATA_DIR / "prices.json"
    meta_file = DATA_DIR / "meta.json"

    with prices_file.open("w") as f:
        json.dump({
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "start_date": START_DATE,
            "products": products_out,
        }, f, separators=(",", ":"))

    with meta_file.open("w") as f:
        json.dump({
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "products": meta_out, "failed": failed,
        }, f, indent=2, ensure_ascii=False)

    print(f"\n{'='*60}")
    print(f"Zapisano {prices_file} ({prices_file.stat().st_size/1024:.1f} KB)")
    print(f"Zapisano {meta_file}")
    print(f"Udane serie: {len(products_out)} | Nieudane: {len(failed)}")
    if failed:
        print(f"Nieudane: {', '.join(failed)}")


if __name__ == "__main__":
    main()
